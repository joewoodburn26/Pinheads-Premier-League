import html
import json
import pathlib
import re
import sys

import openpyxl


def slug(value):
    return re.sub(r"[^a-z0-9]+", "-", str(value).strip().lower()).strip("-")


def number(value, default=0):
    try:
        return int(float(value)) if value not in (None, "") else default
    except Exception:
        return default


def clean(value):
    if value in (None, "", "-"):
        return None
    return str(value).strip()


def main():
    workbook_path = pathlib.Path(sys.argv[1])
    output_path = pathlib.Path(sys.argv[2])
    workbook = openpyxl.load_workbook(workbook_path, data_only=True)

    seasons = [
        {
            "id": "season-2026",
            "name": "Season 2026",
            "draftBudget": 105,
            "activeSeason": True,
            "archived": False,
            "createdAt": "2026-01-01",
        },
        {
            "id": "season-2027",
            "name": "Season 2027",
            "draftBudget": 105,
            "activeSeason": False,
            "archived": False,
            "createdAt": "2026-06-07",
        },
    ]

    pokedex = workbook["Pokedex"]
    pokemon = []
    seen = set()
    for row in range(2, pokedex.max_row + 1):
        dex_number = number(pokedex.cell(row, 1).value, None)
        name = clean(pokedex.cell(row, 2).value)
        if not dex_number or not name:
            continue

        pokemon_id = f"pokemon-{dex_number}-{slug(name)}"
        if pokemon_id in seen:
            continue
        seen.add(pokemon_id)

        hp = number(pokedex.cell(row, 8).value)
        attack = number(pokedex.cell(row, 9).value)
        defense = number(pokedex.cell(row, 10).value)
        special_attack = number(pokedex.cell(row, 11).value)
        special_defense = number(pokedex.cell(row, 12).value)
        speed = number(pokedex.cell(row, 13).value)

        pokemon.append(
            {
                "id": pokemon_id,
                "dexNumber": dex_number,
                "name": name,
                "spriteUrl": f"https://raw.githubusercontent.com/PokeAPI/sprites/master/sprites/pokemon/other/official-artwork/{dex_number}.png",
                "primaryType": clean(pokedex.cell(row, 5).value) or "Normal",
                "secondaryType": clean(pokedex.cell(row, 6).value),
                "hp": hp,
                "attack": attack,
                "defense": defense,
                "specialAttack": special_attack,
                "specialDefense": special_defense,
                "speed": speed,
                "bst": hp + attack + defense + special_attack + special_defense + speed,
                "pointValue": number(pokedex.cell(row, 3).value, 1),
                "legendary": False,
                "mythical": False,
                "paradox": False,
            }
        )

    pokemon_by_name = {entry["name"].lower(): entry for entry in pokemon}

    teams_sheet = workbook["Teams"]
    team_blocks = []
    for header_row in [3, 18]:
        for col in [4, 8, 12, 16]:
            team_name = clean(teams_sheet.cell(header_row, col).value)
            coach_name = clean(teams_sheet.cell(header_row, col + 2).value)
            if team_name and coach_name:
                team_blocks.append((header_row, col, team_name, coach_name))

    coaches = []
    teams = []
    team_pokemon = []
    coach_to_team = {}
    wins = {}
    losses = {}

    for header_row, col, team_name, coach_name in team_blocks:
        coach_id = f"coach-{slug(coach_name)}"
        team_id = f"team-{slug(team_name)}"
        coach_to_team[coach_name.lower()] = team_id
        wins[team_id] = 0
        losses[team_id] = 0
        coaches.append({"id": coach_id, "name": coach_name, "imageUrl": None, "bio": ""})
        teams.append(
            {
                "id": team_id,
                "seasonId": "season-2026",
                "coachId": coach_id,
                "teamName": team_name,
                "logoUrl": None,
                "wins": 0,
                "losses": 0,
            }
        )

        for pick_number, row in enumerate(range(header_row + 2, header_row + 12), start=1):
            pokemon_name = clean(teams_sheet.cell(row, col).value)
            if not pokemon_name:
                continue
            picked = pokemon_by_name[pokemon_name.lower()]
            team_pokemon.append(
                {
                    "id": f"slot-{team_id}-{pick_number}",
                    "seasonId": "season-2026",
                    "teamId": team_id,
                    "pokemonId": picked["id"],
                    "pokemon": picked,
                }
            )

    schedule_sheet = workbook["Schedule"]
    schedule = []
    week = None
    match_number = 0
    for row in range(1, schedule_sheet.max_row + 1):
        values = [schedule_sheet.cell(row, col).value for col in range(1, schedule_sheet.max_column + 1)]
        for value in values:
            if isinstance(value, str) and value.strip().lower().startswith("week"):
                match = re.search(r"(\d+)", value)
                if match:
                    week = int(match.group(1))

        coach_one = clean(schedule_sheet.cell(row, 2).value)
        coach_two = clean(schedule_sheet.cell(row, 10).value)
        if not coach_one or not coach_two:
            continue
        if coach_one.lower() not in coach_to_team or coach_two.lower() not in coach_to_team:
            continue

        status_one = None
        for col in [3, 4]:
            value = clean(schedule_sheet.cell(row, col).value)
            if value in ("W", "L"):
                status_one = value
        status_two = clean(schedule_sheet.cell(row, 9).value)

        home_team = coach_to_team[coach_one.lower()]
        away_team = coach_to_team[coach_two.lower()]
        winner = None
        if status_one == "W":
            winner = home_team
            wins[home_team] += 1
            losses[away_team] += 1
        elif status_two == "W":
            winner = away_team
            wins[away_team] += 1
            losses[home_team] += 1

        replay_links = []
        for value in values:
            if isinstance(value, str) and value.startswith("https://replay.pokemonshowdown.com") and value not in replay_links:
                replay_links.append(value)

        match_number += 1
        schedule.append(
            {
                "id": f"match-2026-{match_number}",
                "seasonId": "season-2026",
                "week": week or 1,
                "homeTeam": home_team,
                "awayTeam": away_team,
                "winner": winner,
                "replay1": replay_links[0] if len(replay_links) > 0 else None,
                "replay2": replay_links[1] if len(replay_links) > 1 else None,
                "replay3": replay_links[2] if len(replay_links) > 2 else None,
            }
        )

    for team in teams:
        team["wins"] = wins[team["id"]]
        team["losses"] = losses[team["id"]]

    stats = []
    for slot in team_pokemon:
        stats.append(
            {
                "id": f"stat-{slot['teamId']}-{slot['pokemonId']}",
                "seasonId": "season-2026",
                "pokemonId": slot["pokemonId"],
                "teamId": slot["teamId"],
                "gamesPlayed": 0,
                "wins": 0,
                "losses": 0,
                "kos": 0,
                "deaths": 0,
                "pokemon": slot["pokemon"],
            }
        )

    rules = []
    banned = []
    rules_sheet = workbook["Rules"]
    for row in range(1, rules_sheet.max_row + 1):
        rule_number = rules_sheet.cell(row, 2).value
        rule_text = clean(rules_sheet.cell(row, 3).value)
        banned_name = clean(rules_sheet.cell(row, 5).value)
        if rule_text and isinstance(rule_number, (int, float)) and row < 26:
            rules.append(rule_text)
        if rule_text and banned_name and isinstance(rule_number, (int, float)) and row >= 26:
            banned.append(f"{rule_text}: {banned_name}")

    rules_html = (
        "<h2>Draft Rules</h2><ol>"
        + "".join(f"<li>{html.escape(rule)}</li>" for rule in rules)
        + "</ol><h2>Banned Abilities, Items & Moves</h2><ul>"
        + "".join(f"<li>{html.escape(rule)}</li>" for rule in banned)
        + "</ul>"
    )

    output = 'import type { Coach, Pokemon, PokemonStats, ScheduleMatch, Season, Team, TeamPokemon } from "@/lib/types";\n\n'
    exports = {
        "seasons": ("Season[]", seasons),
        "coaches": ("Coach[]", coaches),
        "teams": ("Team[]", teams),
        "pokemon": ("Pokemon[]", pokemon),
        "teamPokemon": ("TeamPokemon[]", team_pokemon),
        "schedule": ("ScheduleMatch[]", schedule),
        "stats": ("PokemonStats[]", stats),
    }
    for name, (type_name, value) in exports.items():
        output += f"export const {name}: {type_name} = {json.dumps(value, ensure_ascii=True, indent=2)};\n\n"
    output += f"export const rulesHtml = {json.dumps(rules_html, ensure_ascii=True)};\n"

    output_path.write_text(output, encoding="utf-8")
    print(
        json.dumps(
            {
                "pokemon": len(pokemon),
                "teams": len(teams),
                "rosterSlots": len(team_pokemon),
                "matches": len(schedule),
                "rules": len(rules),
                "banned": len(banned),
                "standings": [(team["teamName"], team["wins"], team["losses"]) for team in teams],
            },
            indent=2,
        )
    )


if __name__ == "__main__":
    main()
